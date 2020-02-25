# -*- coding: utf-8 -*-
import os
from openpyxl.reader.excel import load_workbook
import shutil
import re
from datetime import date

patt = re.compile('\d\d\d\d-\d\d-\d\d')

folder = '/home/gntech/Documents/Familia/Juliana/'
filename = 'Calendario_COC.xlsx'
path = os.path.join(folder,filename)

today = str(date.today())

#shutil.copyfile(path, path.replace('.xlsx', 'copy.xlsx'))

wb = load_workbook(path)

ws_base = wb['Agendamento_Marco']

dic = {}
cell_name = ws_base.cell(column=1, row=4).value
row = 4
while bool(cell_name):
    cell_name = ws_base.cell(column=1, row=row).value
    cell_date = ws_base.cell(column=2, row=row).value
    cell_city = ws_base.cell(column=4, row=row).value
    if cell_name in dic.keys():
        dic[cell_name][cell_date] = cell_city
    else:
        dic[cell_name] = {cell_date: cell_city}
    row +=1

ws_marco = wb['Planilha1']
row = 3
cell = True
while bool(cell):
    data = ws_marco.cell(column=1, row=row).value
    nome = True
    col = 3
    while bool(nome):
        nome = ws_marco.cell(column=col, row=2).value
        if nome in dic.keys():
            if data in dic[nome].keys():
                r = dic[nome][data]
                ws_marco.cell(row=row, column=col).value = r
        col += 1
        if col > 1000:break
    row += 1
    if row > 1000:break
wb.save(path.replace('.xlsx', 'copy.xlsx'))









